"""Upload test cases from an Excel file to Azure DevOps.

Reads test cases from any Excel file that follows the standard column format
and uploads them as Test Case work items, linked to their parent User Stories.

Expected Excel columns (case-insensitive, order doesn't matter):
  - Test Case ID     : e.g. TC-UC20-001  (used for display/logging only)
  - Module           : e.g. "Future Date Handling"
  - Feature          : e.g. "Scan Receipt"
  - Scenario         : becomes the work item Title
  - Preconditions    : free text
  - Test Steps       : numbered steps, one per line  e.g. "1. Do X\n2. Do Y"
  - Expected Result  : single overall expected result string
  - Priority         : High / Medium / Low / Critical
  - Type             : Functional / Negative / Edge / Regression / UI
  - Story IDs        : comma-separated VSTS IDs  e.g. "36735, 36736"
  - Platform         : Web / Mobile / Web & Mobile / Mobile only / etc.
  - UC               : optional label e.g. "UC20"  (grouping only)

Steps + Expected Result handling:
  Format A -- numbered steps + single expected result:
    "1. Do X\n2. Do Y" → expected attached to last step.
  Format B -- pipe-separated per-step expected:
    "1. Do X | Result X\n2. Do Y | Result Y" → each line is its own pair.

Usage:
  python upload_test_cases_from_excel.py path/to/test_cases.xlsx
  python upload_test_cases_from_excel.py path/to/test_cases.xlsx --dry-run
  python upload_test_cases_from_excel.py path/to/test_cases.xlsx --sheet "Sheet1"
"""
from __future__ import annotations

import html
import os
import sys
import time
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
import openpyxl

load_dotenv()

# Windows consoles default to cp1252, which can't encode characters like
# '→' that show up in test-step text -- force UTF-8 so printing never crashes.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

# -- Constants ------------------------------------------------------------------

ORG_URL  = os.getenv("ORGANIZATION_URL", "https://dev.azure.com/expenseondemand")
PAT      = os.getenv("AZURE_DEVOPS_PAT")
PROJECT  = os.getenv("PROJECT_NAME", "Solo Expenses")
DRY_RUN  = "--dry-run" in sys.argv

AREA_PATH         = "Solo Expenses\\Expense On Demand"
AUTOMATION_VAL    = "To be automated"
PRECONDITION_FIELD = "EODAgile.Precondition"

_PRIORITY_MAP = {"critical": 1, "high": 2, "medium": 3, "low": 4}

_AREA_MAP = {
    "web & mobile":           "Web & Mobile",
    "mobile only":            "Mobile (iOS & Android)",
    "mobile – android & ios": "Mobile (iOS & Android)",
    "mobile -- android & ios": "Mobile (iOS & Android)",
    "mobile":                 "Mobile (iOS & Android)",
    "android & ios":          "Mobile (iOS & Android)",
    "android":                "Android",
    "ios":                    "iOS",
    "api":                    "API",
    "web":                    "Web",
}

# Full allowed list from ADO (used for validation in the prompt).
_ALLOWED_FUNCTIONS = [
    "Accruals Report", "Additional Fields", "Administrator Priviliges",
    "Advances", "AI Function", "Analytics", "Android - SDK Change",
    "Approver", "Approver Historical View", "AppStore",
    "Assign Category to Grade", "Assisted Support", "Attendance",
    "Audit trail", "Auto Delete Old & Unclaimed Expenses", "Automation",
    "Billing Plan", "Business Central", "Business Purpose",
    "Carbon Footprint", "Change Password", "Chatbot", "CICD", "Claimant",
    "Compliance Statement", "Configure Approval", "Configure Culumns",
    "Container", "Cost Analysis", "Cost Centre", "Counter Approver",
    "Create Expense", "Credit Card", "Customer", "Dashboard", "Delegate",
    "Delete Data", "Department", "Deputy", "DOC", "Emails",
    "Expense Category Management", "Expense List", "Expense Policy",
    "Export", "FC/Cash Allowance", "Finance Approver", "Foreign Exchange",
    "Grades", "identity Server", "Import", "Integrations",
    "Laundry Policy", "Leave", "Location", "Login", "Manage Employee",
    "Marketing Panels", "MFA", "Mileage", "Multi City", "Multi Company",
    "Multi Country", "My Organisation", "My Profile", "My Subscription",
    "NetSuite", "New Design", "Notifications", "Open API", "Optimisation",
    "PDF", "Pre-Approval", "Pricing Calculator", "Projects & Budget",
    "Quick Setup", "QuickBooks", "Quickbooks Desktop",
    "Receipt with multiple TAX", "Refer & Earn", "Registration",
    "Report Writer", "Sage 200", "Sage 50", "Search",
    "Single Sign On (SSO)", "Smart Scan", "Split Entertainment",
    "Split Expense Category", "Stripe", "Supplier", "Support", "Tally",
    "Tax", "Time", "Translator", "Travel", "Universal Import", "UPI",
    "VAT No VAT/ No Receipt", "Vendor / AP automation", "Vendors",
    "Video", "Voice Activation", "Website", "Weekend Policy",
    "Workflow", "Xero", "YBL - Claimant", "Your Organisation",
]

# Shortcuts shown in the numbered menu (most commonly used).
_COMMON_FUNCTIONS = [
    "Smart Scan", "Credit Card", "Create Expense", "Expense Policy",
    "Import", "Export", "Dashboard", "Approver", "Mileage", "Travel",
]

_AREA_OPTIONS = [
    "Web & Mobile", "Web", "Mobile (iOS & Android)",
    "Android", "iOS", "API",
]

# 2026 iteration paths as (label, full_path) -- current month highlighted.
_ITERATIONS_2026 = [
    ("June 2026  <- current",   r"Solo Expenses\2026\Iteration 2026 06-June"),
    ("July 2026",               r"Solo Expenses\2026\Iteration 2026 07-July"),
    ("August 2026",             r"Solo Expenses\2026\Iteration 2026 08-August"),
    ("September 2026",          r"Solo Expenses\2026\Iteration 2026 09-September"),
    ("October 2026",            r"Solo Expenses\2026\Iteration 2026 10-October"),
    ("November 2026",           r"Solo Expenses\2026\Iteration 2026 11-November"),
    ("December 2026",           r"Solo Expenses\2026\Iteration 2026 12-December"),
    ("May 2026",                r"Solo Expenses\2026\Iteration 2026 05-May"),
    ("April 2026",              r"Solo Expenses\2026\Iteration 2026 04-April"),
    ("Backlog",                 r"Solo Expenses\Backlog"),
]


def _priority_int(s: str) -> int:
    return _PRIORITY_MAP.get(s.lower().strip(), 3)


def _custom_area(platform: str, fallback: str = "Web & Mobile") -> str:
    key = platform.lower().strip()
    for fragment, value in _AREA_MAP.items():
        if fragment in key:
            return value
    return fallback


# -- Interactive setup wizard ---------------------------------------------------

def _pick(prompt: str, options: list[str], allow_custom: bool = False) -> str:
    """Show a numbered list and return the chosen value."""
    print()
    for i, opt in enumerate(options, 1):
        print(f"    {i:>2})  {opt}")
    if allow_custom:
        print(f"         (or type any other value)")
    while True:
        raw = input(f"\n  {prompt}: ").strip()
        if raw.isdigit():
            idx = int(raw) - 1
            if 0 <= idx < len(options):
                # Return just the path (second element) for iteration tuples,
                # or the string itself for simple lists.
                return options[idx]
        elif raw:
            if allow_custom:
                return raw
            # Try case-insensitive match
            match = next((o for o in options if o.lower() == raw.lower()), None)
            if match:
                return match
        print("  Please enter a number from the list" +
              (" or type a custom value" if allow_custom else "") + ".")


def _interactive_setup(test_cases: list[dict]) -> dict:
    """
    Walk the user through the four required upload settings.
    Returns a config dict: function, iteration, area, global_story_ids.
    """
    W = 60
    print()
    print("-" * W)
    print("  ADO Test Case Upload -- Setup")
    print("-" * W)
    print(f"  {len(test_cases)} test case(s) loaded")
    print(f"  Mode : {'DRY RUN' if DRY_RUN else 'LIVE UPLOAD'}")

    # -- [1/4] Function ---------------------------------------------------------
    print()
    print("  [1/4]  Function  (Custom.function -- required field in ADO)")
    print("         Common values:")
    chosen_function = _pick("Enter number or type exact name",
                            _COMMON_FUNCTIONS, allow_custom=True)
    # Validate against allowed list (warn but don't block -- org list may grow).
    if chosen_function not in _ALLOWED_FUNCTIONS:
        print(f"\n  ⚠  '{chosen_function}' is not in the confirmed ADO list.")
        print("     Upload will fail if this value doesn't match the dropdown exactly.")
        confirm = input("  Continue anyway? [y/N] ").strip().lower()
        if confirm != "y":
            print("  Aborted -- re-run and pick a valid function.")
            sys.exit(0)

    # -- [2/4] Iteration --------------------------------------------------------
    print()
    print("  [2/4]  Iteration Path  (which sprint to assign these TCs to)")
    print("         Leave unset = items go to root iteration (off sprint board)")
    print()
    iter_labels  = [label for label, _ in _ITERATIONS_2026]
    iter_labels.append("Skip -- don't set iteration")
    chosen_label = _pick("Enter number", iter_labels, allow_custom=False)

    if chosen_label == "Skip -- don't set iteration":
        chosen_iteration = None
    else:
        # Find the matching path
        chosen_iteration = next(
            (path for label, path in _ITERATIONS_2026 if label == chosen_label), None
        )

    # -- [3/4] Area (Custom.Area) -----------------------------------------------
    print()
    print("  [3/4]  Default Area  (Custom.Area -- applies where Platform column")
    print("         is blank or unrecognised; per-row Platform column still overrides)")
    chosen_area = _pick("Enter number", _AREA_OPTIONS, allow_custom=False)

    # -- [4/4] Story IDs --------------------------------------------------------
    print()
    print("  [4/4]  Story IDs")
    rows_with_ids    = sum(1 for tc in test_cases if tc["story_ids"])
    rows_without_ids = len(test_cases) - rows_with_ids
    print(f"         {rows_with_ids} row(s) already have Story IDs in the Excel.")
    if rows_without_ids:
        print(f"         {rows_without_ids} row(s) have none -- enter global IDs below")
        print("         to apply them to those rows, or press Enter to skip.")
    else:
        print("         All rows have Story IDs -- press Enter to skip global override.")

    raw_ids = input("\n  Story IDs (comma-separated) or Enter to skip: ").strip()
    global_story_ids: list[int] = []
    for s in raw_ids.replace(";", ",").split(","):
        s = s.strip()
        if s.isdigit():
            global_story_ids.append(int(s))

    # -- Summary ----------------------------------------------------------------
    print()
    print("-" * W)
    print("  Ready to upload:")
    print(f"    Function  : {chosen_function}")
    print(f"    Iteration : {chosen_iteration or '(not set)'}")
    print(f"    Area      : {chosen_area}  (default -- Platform column overrides per row)")
    if global_story_ids:
        print(f"    Story IDs : {global_story_ids}  (applied to rows without their own)")
    else:
        print(f"    Story IDs : from Excel per row")
    print(f"    Cases     : {len(test_cases)}")
    print(f"    Mode      : {'DRY RUN (no changes made)' if DRY_RUN else 'LIVE UPLOAD'}")
    print("-" * W)

    if not DRY_RUN:
        confirm = input("\n  Proceed with live upload? [y/N] ").strip().lower()
        if confirm != "y":
            print("  Aborted.")
            sys.exit(0)

    return {
        "function":         chosen_function,
        "iteration":        chosen_iteration,
        "area":             chosen_area,
        "global_story_ids": global_story_ids,
    }


def _apply_config(test_cases: list[dict], config: dict) -> None:
    """
    Apply the wizard config as defaults onto each test case (in-place).
    Per-row Excel values take priority over the global config.
    """
    for tc in test_cases:
        # Function: Excel column wins; wizard is the fallback.
        if not tc.get("function"):
            tc["function"] = config["function"]

        # Area: only override if the Platform column didn't resolve to a known area.
        # We store the raw platform string and resolve at patch time.
        tc["_area_fallback"] = config["area"]

        # Story IDs: append global IDs to rows that have none.
        if not tc["story_ids"] and config["global_story_ids"]:
            tc["story_ids"] = config["global_story_ids"]


# -- Excel reading --------------------------------------------------------------

def _normalise_header(h: str) -> str:
    return " ".join(str(h).lower().split()) if h else ""


def load_test_cases_from_excel(path: str, sheet_name: Optional[str] = None) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        selected = wb.active
        for name in wb.sheetnames:
            nl = name.lower()
            if "summary" not in nl and "coverage" not in nl and "readme" not in nl:
                selected = wb[name]
                break
        ws = selected
        print(f"  Reading sheet: '{ws.title}'")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sheet appears to be empty.")

    col_map: dict[str, int] = {}
    for i, h in enumerate(rows[0]):
        if h is not None:
            col_map[_normalise_header(str(h))] = i

    if "scenario" not in col_map:
        raise ValueError(
            f"Required column 'scenario' not found. Found: {list(col_map.keys())}"
        )

    def _get(row: tuple, key: str, default: str = "") -> str:
        idx = col_map.get(key)
        if idx is None:
            return default
        val = row[idx]
        return str(val).strip() if val is not None else default

    test_cases: list[dict] = []
    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        scenario = _get(row, "scenario")
        if not scenario:
            continue

        story_ids: list[int] = []
        for s in _get(row, "story ids").replace(";", ",").split(","):
            s = s.strip()
            if s.isdigit():
                story_ids.append(int(s))

        steps_raw = _get(row, "test steps")
        expected  = _get(row, "expected result")

        tc = {
            "row":           row_num,
            "tc_id":         _get(row, "test case id"),
            "uc":            _get(row, "uc"),
            "module":        _get(row, "module"),
            "feature":       _get(row, "feature"),
            "title":         scenario,
            "preconditions": _get(row, "preconditions"),
            "steps":         _parse_steps(steps_raw, expected),
            "test_data":     _get(row, "test data"),
            "priority":      _priority_int(_get(row, "priority", "medium")),
            "type":          _get(row, "type", "Functional"),
            "platform":      _get(row, "platform", ""),
            "story_ids":     story_ids,
            "function":      _get(row, "function"),  # empty = wizard fills in
            "_area_fallback": "Web & Mobile",         # replaced by wizard
        }
        test_cases.append(tc)

    print(f"  Loaded {len(test_cases)} test cases from '{ws.title}'")
    return test_cases


def _parse_steps(steps_raw: str, overall_expected: str) -> list[tuple[str, str]]:
    if not steps_raw:
        return [("(no steps provided)", overall_expected or "")]

    lines = [l.strip() for l in steps_raw.splitlines() if l.strip()]

    if any(" | " in line for line in lines):
        pairs: list[tuple[str, str]] = []
        for line in lines:
            if " | " in line:
                action, _, exp = line.partition(" | ")
                pairs.append((_strip_step_num(action), exp.strip()))
            else:
                pairs.append((_strip_step_num(line), ""))
        return pairs

    pairs = [(_strip_step_num(line), "") for line in lines]
    if overall_expected and pairs:
        last_action, _ = pairs[-1]
        pairs[-1] = (last_action, overall_expected)
    elif overall_expected:
        pairs.append(("Verify outcome", overall_expected))
    return pairs


def _strip_step_num(text: str) -> str:
    import re
    return re.sub(r"^\d+[\.\)]\s*", "", text.strip())


# -- ADO helpers ----------------------------------------------------------------

def _steps_xml(steps: list[tuple[str, str]]) -> str:
    def enc(t: str) -> str:
        return html.escape(f"<DIV><P>{t}</P></DIV>")

    parts = [f'<steps id="0" last="{len(steps)}">']
    for i, (action, expected) in enumerate(steps, 1):
        step_type = "ValidateStep" if expected else "ActionStep"
        parts.append(
            f'<step id="{i}" type="{step_type}">'
            f'<parameterizedString isformatted="true">{enc(action)}</parameterizedString>'
            f'<parameterizedString isformatted="true">{enc(expected)}</parameterizedString>'
            f'<description/></step>'
        )
    parts.append("</steps>")
    return "".join(parts)


def _make_patch(tc: dict, iteration: Optional[str]) -> list:
    from azure.devops.v7_1.work_item_tracking.models import JsonPatchOperation

    title = f"{tc['tc_id']}: {tc['title']}" if tc.get("tc_id") else tc["title"]
    area  = _custom_area(tc["platform"], fallback=tc.get("_area_fallback", "Web & Mobile"))

    ops = [
        JsonPatchOperation(op="add", path="/fields/System.Title",                   value=title),
        JsonPatchOperation(op="add", path="/fields/System.AreaPath",                value=AREA_PATH),
        JsonPatchOperation(op="add", path="/fields/System.State",                   value="Design"),
        JsonPatchOperation(op="add", path="/fields/Microsoft.VSTS.Common.Priority", value=tc["priority"]),
        JsonPatchOperation(op="add", path="/fields/Microsoft.VSTS.TCM.Steps",       value=_steps_xml(tc["steps"])),
        JsonPatchOperation(op="add", path="/fields/Custom.function",                value=tc["function"]),
        JsonPatchOperation(op="add", path="/fields/Custom.Area",                    value=area),
        JsonPatchOperation(op="add", path="/fields/Custom.Automation",              value=AUTOMATION_VAL),
    ]

    if iteration:
        ops.append(JsonPatchOperation(
            op="add", path="/fields/System.IterationPath", value=iteration,
        ))

    if tc.get("preconditions"):
        ops.append(JsonPatchOperation(
            op="add", path=f"/fields/{PRECONDITION_FIELD}", value=tc["preconditions"],
        ))

    if tc.get("test_data"):
        ops.append(JsonPatchOperation(
            op="add", path="/fields/Microsoft.VSTS.TCM.LocalDataSource", value=tc["test_data"],
        ))

    for sid in tc.get("story_ids", []):
        ops.append(JsonPatchOperation(
            op="add", path="/relations/-",
            value={
                "rel": "Microsoft.VSTS.Common.TestedBy-Reverse",
                "url": f"{ORG_URL}/_apis/wit/workItems/{sid}",
                "attributes": {"comment": f"Test case for story {sid}"},
            },
        ))

    return ops


# -- Upload ---------------------------------------------------------------------

def upload(test_cases: list[dict], config: dict, wit_client) -> None:
    iteration = config["iteration"]

    groups: dict[str, list[dict]] = {}
    for tc in test_cases:
        label = tc.get("uc") or tc.get("module") or "General"
        groups.setdefault(label, []).append(tc)

    total = created = failed = skipped = 0

    for group_label, tcs in groups.items():
        print(f"\n-- {group_label} ({len(tcs)} test cases) --")
        for tc in tcs:
            total += 1
            display = f"{tc.get('tc_id', '')} {tc['title']}".strip()

            if not tc["story_ids"]:
                print(f"  [SKIP] {display}  -- no Story IDs")
                skipped += 1
                continue

            if DRY_RUN:
                area = _custom_area(tc["platform"], fallback=tc.get("_area_fallback", "Web & Mobile"))
                print(f"  [DRY]  {display}")
                print(f"         fn={tc['function']}  area={area}  "
                      f"pri={tc['priority']}  stories={tc['story_ids']}  steps={len(tc['steps'])}")
                created += 1
                continue

            try:
                wi = wit_client.create_work_item(
                    document=_make_patch(tc, iteration),
                    project=PROJECT,
                    type="Test Case",
                )
                print(f"  [OK]   #{wi.id} -- {display}")
                created += 1
            except Exception as e:
                print(f"  [FAIL] {display}\n         {e}")
                failed += 1

            time.sleep(0.25)

    print(f"\n{'-' * 55}")
    print(f"{'DRY RUN -- ' if DRY_RUN else ''}Results: "
          f"{created} created, {failed} failed, {skipped} skipped / {total} total")


# -- Entry point ----------------------------------------------------------------

def _get_arg(flag: str) -> Optional[str]:
    for i, arg in enumerate(sys.argv):
        if arg.startswith(f"{flag}="):
            return arg.split("=", 1)[1]
        if arg == flag and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return None


if __name__ == "__main__":
    excel_path = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
    if not excel_path:
        print("Usage: python upload_test_cases_from_excel.py <path.xlsx> [--dry-run] [--sheet SheetName]")
        sys.exit(1)

    if not Path(excel_path).exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    if not PAT and not DRY_RUN:
        print("ERROR: AZURE_DEVOPS_PAT not set in .env")
        sys.exit(1)

    print(f"\n  File    : {excel_path}")
    print(f"  Project : {PROJECT}  |  Org : {ORG_URL}")
    print()

    try:
        test_cases = load_test_cases_from_excel(excel_path, _get_arg("--sheet"))
    except Exception as e:
        print(f"ERROR reading Excel: {e}")
        sys.exit(1)

    if not test_cases:
        print("No test cases found -- nothing to upload.")
        sys.exit(0)

    # Interactive setup -- asks for function, iteration, area, story IDs.
    config = _interactive_setup(test_cases)

    # Apply wizard choices as defaults onto each TC.
    _apply_config(test_cases, config)

    if DRY_RUN:
        upload(test_cases, config, wit_client=None)
    else:
        from azure.devops.connection import Connection
        from msrest.authentication import BasicAuthentication
        creds = BasicAuthentication("", PAT)
        conn  = Connection(base_url=ORG_URL, creds=creds)
        upload(test_cases, config, conn.clients.get_work_item_tracking_client())
